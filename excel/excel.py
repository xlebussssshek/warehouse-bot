import openpyxl
from openpyxl import Workbook
from datetime import datetime


def create_stock_report(data):
    filename = f"Остатки_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"

    headers = ["Артикул", "Наименование", "Количество"]
    ws.append(headers)

    for row in data:
        ws.append([row[0], row[1], row[2]])

    for col in range(1, 4):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    wb.save(filename)
    return filename

def history_report(data):
    filename = f"История_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "История"

    headers = ["id", "Артикул", "Тип операции", "Количество", "Прошлое количество", "Новое количество", "user_id", "Детали", "Время операции"]
    ws.append(headers)

    for row in data:
        ws.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8]])

    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    wb.save(filename)
    return filename